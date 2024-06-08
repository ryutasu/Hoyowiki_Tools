import json
import os
import string
import time
import requests
import openpyxl
from openpyxl.styles import *

file_urls = {
    'CHS': 'https://raw.githubusercontent.com/Dimbreath/StarRailData/master/TextMap/TextMapCHS.json',
    'CHT': 'https://raw.githubusercontent.com/Dimbreath/StarRailData/master/TextMap/TextMapCHT.json',
    'DE': 'https://raw.githubusercontent.com/Dimbreath/StarRailData/master/TextMap/TextMapDE.json',
    'EN': 'https://raw.githubusercontent.com/Dimbreath/StarRailData/master/TextMap/TextMapEN.json',
    'ES': 'https://raw.githubusercontent.com/Dimbreath/StarRailData/master/TextMap/TextMapES.json',
    'FR': 'https://raw.githubusercontent.com/Dimbreath/StarRailData/master/TextMap/TextMapFR.json',
    'ID': 'https://raw.githubusercontent.com/Dimbreath/StarRailData/master/TextMap/TextMapID.json',
    'JP': 'https://raw.githubusercontent.com/Dimbreath/StarRailData/master/TextMap/TextMapJP.json',
    'KR': 'https://raw.githubusercontent.com/Dimbreath/StarRailData/master/TextMap/TextMapKR.json',
    'PT': 'https://raw.githubusercontent.com/Dimbreath/StarRailData/master/TextMap/TextMapPT.json',
    'RU': 'https://raw.githubusercontent.com/Dimbreath/StarRailData/master/TextMap/TextMapRU.json',
    'TH': 'https://raw.githubusercontent.com/Dimbreath/StarRailData/master/TextMap/TextMapTH.json',
    'VI': 'https://raw.githubusercontent.com/Dimbreath/StarRailData/master/TextMap/TextMapVI.json'
}
Achievement_urls = {
    #'TextJoinItem': 'https://raw.githubusercontent.com/Dimbreath/StarRailData/master/ExcelOutput/TextJoinItem.json',
    #'TextJoinConfig': 'https://raw.githubusercontent.com/Dimbreath/StarRailData/master/ExcelOutput/TextJoinConfig.json'
    'AchievementData': 'https://raw.githubusercontent.com/Dimbreath/StarRailData/master/ExcelOutput/AchievementData.json',
    'AchievementSeries': 'https://raw.githubusercontent.com/Dimbreath/StarRailData/master/ExcelOutput/AchievementSeries.json'
}

def filling(start_loc,end_loc,ws):         #参数为左上角坐标和右下角坐标，形如'D3','A5'等。ws是worksheet对象。
    x_start = start_loc[0]
    y_start = start_loc[1:len(start_loc)]   #切片获取坐标的数字部分
    x_end = end_loc[0]
    y_end = end_loc[1:len(end_loc)]   
    len_y = int(y_end) - int(y_start) + 1
    alphabet = string.ascii_uppercase        #导入字母表
    len_x = alphabet.index(x_end) - alphabet.index(x_start) + 1
    # 左上
    temp = start_loc
    ws[temp].border = Border(left=Side(style='thick'),top=Side(style='thick'))
    # 右下
    temp = end_loc
    ws[temp].border = Border(right=Side(style='thick'),bottom=Side(style='thick'))
    # 右上
    temp = x_end + y_start
    ws[temp].border = Border(right=Side(style='thick'),top=Side(style='thick'))
    # 左下
    temp = x_start + y_end
    ws[temp].border = Border(left=Side(style='thick'),bottom=Side(style='thick'))
    # 上
    for i in range(0,len_x-2):
        temp = alphabet[alphabet.index(x_start)+1+i] + y_start
        ws[temp].border = Border(top=Side(style='thick'))
    # 下
    for i in range(0,len_x-2):
        temp = alphabet[alphabet.index(x_start)+1+i] + y_end
        ws[temp].border = Border(bottom=Side(style='thick'))
    # 左
    for i in range(0,len_y-2):
        temp = x_start + str(int(y_start) + 1 + i)
        ws[temp].border = Border(left=Side(style='thick'))
    # 右
    for i in range(0,len_y-2):
        temp = x_end + str(int(y_start) + 1 + i)
        ws[temp].border = Border(right=Side(style='thick'))
    return 0

def checkRarity(Rarity):
    if Rarity == "Low":
        return 5
    elif Rarity == "Mid":
        return 10
    elif Rarity == "High":
        return 20

def getimage(hash_value,price):
    imagelist = {
        'A1':'https://iotapp.ttu.edu.tw/ionic-project/Ryutasu3/CultivateAchievementIcon3.png',
        'A2':'https://iotapp.ttu.edu.tw/ionic-project/Ryutasu3/CultivateAchievementIcon2.png',
        'A3':'https://iotapp.ttu.edu.tw/ionic-project/Ryutasu3/CultivateAchievementIcon1.png',
        'B1':'https://iotapp.ttu.edu.tw/ionic-project/Ryutasu3/CollectAchievementIcon3.png',
        'B2':'https://iotapp.ttu.edu.tw/ionic-project/Ryutasu3/CollectAchievementIcon2.png',
        'B3':'https://iotapp.ttu.edu.tw/ionic-project/Ryutasu3/CollectAchievementIcon1.png',
        'C1':'https://iotapp.ttu.edu.tw/ionic-project/Ryutasu3/MainLineAchievementIcon3.png',
        'C2':'https://iotapp.ttu.edu.tw/ionic-project/Ryutasu3/MainLineAchievementIcon2.png',
        'C3':'https://iotapp.ttu.edu.tw/ionic-project/Ryutasu3/MainLineAchievementIcon1.png',
        'D1':'https://iotapp.ttu.edu.tw/ionic-project/Ryutasu3/ExploreAchievementIcon3.png',
        'D2':'https://iotapp.ttu.edu.tw/ionic-project/Ryutasu3/ExploreAchievementIcon2.png',
        'D3':'https://iotapp.ttu.edu.tw/ionic-project/Ryutasu3/ExploreAchievementIcon1.png',
        'E1':'https://iotapp.ttu.edu.tw/ionic-project/Ryutasu3/ReputationAchievementIcon3.png',
        'E2':'https://iotapp.ttu.edu.tw/ionic-project/Ryutasu3/ReputationAchievementIcon2.png',
        'E3':'https://iotapp.ttu.edu.tw/ionic-project/Ryutasu3/ReputationAchievementIcon1.png',
        'F1':'https://iotapp.ttu.edu.tw/ionic-project/Ryutasu3/ChallengeAchievementIcon3.png',
        'F2':'https://iotapp.ttu.edu.tw/ionic-project/Ryutasu3/ChallengeAchievementIcon2.png',
        'F3':'https://iotapp.ttu.edu.tw/ionic-project/Ryutasu3/ChallengeAchievementIcon1.png',
        'G1':'https://iotapp.ttu.edu.tw/ionic-project/Ryutasu3/BattleAchievementIcon3.png',
        'G2':'https://iotapp.ttu.edu.tw/ionic-project/Ryutasu3/BattleAchievementIcon2.png',
        'G3':'https://iotapp.ttu.edu.tw/ionic-project/Ryutasu3/BattleAchievementIcon1.png',
        'H1':'https://iotapp.ttu.edu.tw/ionic-project/Ryutasu3/TrackAchievementIcon3.png',
        'H2':'https://iotapp.ttu.edu.tw/ionic-project/Ryutasu3/TrackAchievementIcon2.png',
        'H3':'https://iotapp.ttu.edu.tw/ionic-project/Ryutasu3/TrackAchievementIcon1.png',
        'I1':'https://iotapp.ttu.edu.tw/ionic-project/Ryutasu3/RougeAchievementIcon3.png',
        'I2':'https://iotapp.ttu.edu.tw/ionic-project/Ryutasu3/RougeAchievementIcon2.png',
        'I3':'https://iotapp.ttu.edu.tw/ionic-project/Ryutasu3/RougeAchievementIcon1.png',
    }
    if(hash_value == -621359642):
        if(price == 5):
            return imagelist['A1']
        elif(price == 10):
            return imagelist['A2']
        elif(price == 20):
            return imagelist['A3']
    elif(hash_value == -621359643):
        if(price == 5):
            return imagelist['B1']
        elif(price == 10):
            return imagelist['B2']
        elif(price == 20):
            return imagelist['B3']
    elif(hash_value == -621359644):
        if(price == 5):
            return imagelist['C1']
        elif(price == 10):
            return imagelist['C2']
        elif(price == 20):
            return imagelist['C3']
    elif(hash_value == -621359637):
        if(price == 5):
            return imagelist['D1']
        elif(price == 10):
            return imagelist['D2']
        elif(price == 20):
            return imagelist['D3']
    elif(hash_value == -621359638):
        if(price == 5):
            return imagelist['E1']
        elif(price == 10):
            return imagelist['E2']
        elif(price == 20):
            return imagelist['E3']
    elif(hash_value == -621359639):
        if(price == 5):
            return imagelist['F1']
        elif(price == 10):
            return imagelist['F2']
        elif(price == 20):
            return imagelist['F3']
    elif(hash_value == -621359640):
        if(price == 5):
            return imagelist['G1']
        elif(price == 10):
            return imagelist['G2']
        elif(price == 20):
            return imagelist['G3']
    elif(hash_value == -621359633):
        if(price == 5):
            return imagelist['H1']
        elif(price == 10):
            return imagelist['H2']
        elif(price == 20):
            return imagelist['H3']
    elif(hash_value == -621359634):
        if(price == 5):
            return imagelist['I1']
        elif(price == 10):
            return imagelist['I2']
        elif(price == 20):
            return imagelist['I3']
    else:
        return 'error'

def output_excel2(excel_output):
    filename2 = "output_product.xlsx"
    IDcount = 1
    showtypecount = 0
    workbook2 = openpyxl.Workbook()
    sheet = workbook2.active
    sheet.title = "product"
    sheet["A1"] = "id"

    sheet[f"A{IDcount}"] = "id"
    sheet[f"B{IDcount}"] = "name"
    sheet[f"C{IDcount}"] = "description"
    sheet[f"D{IDcount}"] = "price"
    sheet[f"E{IDcount}"] = "type"
    sheet[f"F{IDcount}"] = "showtype"
    sheet[f"G{IDcount}"] = "image"
    sheet[f"H{IDcount}"] = "isselect"
    IDcount+=1

    for key,data in excel_output.items():
        hash_value = data["hash_value"]
        for showtype,sdata in data.items():
            for name,adata in sdata.items():
                pdesc = adata["Achievement_Desc"]
                pprice = checkRarity(adata["Rarity"])
                sheet[f"A{IDcount}"] = IDcount-1
                sheet[f"B{IDcount}"] = name
                sheet[f"C{IDcount}"] = pdesc
                sheet[f"D{IDcount}"] = pprice
                sheet[f"E{IDcount}"] = key
                sheet[f"F{IDcount}"] = showtype
                sheet[f"G{IDcount}"] = getimage(hash_value,pprice)
                sheet[f"H{IDcount}"] = "未完成"
                IDcount+=1
            showtypecount+=1
            if(showtypecount == 2):
                showtypecount = 0
                break

    workbook2.save(filename2)
    os.system(f"start excel {filename2}")

def output_excel(excel_output):
    filename = "output.xlsx"
    new_filename = filename
    count = 0

    # 检查文件是否已存在，如果已存在，则生成新的文件名
    while os.path.exists(new_filename):
        count += 1
        new_filename = f"output({count}).xlsx"

    workbook = openpyxl.Workbook()
    alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

    # 遍历excel_output中的每个键，为每个键创建一个工作表
    for key in excel_output:
        worksheet = workbook.create_sheet(title=key)

        for col_letter in ['C','F','K','L']:
            worksheet.column_dimensions[col_letter].width = 32

        for col_letter in ['G', 'H']:
            worksheet.column_dimensions[col_letter].width = 32

        excel_output_shown_data = excel_output[key]["shown"]
        excel_output_hidden_data = excel_output[key]["hidden"]

        excel_count = 5
        total_amount = 0
        total_Stellar_Jade = 0
        shown_cell_filling_start = 'F'
        shown_cell_filling_start += str(excel_count)
        composite_str_filling_start = 'K'
        composite_str_filling_start += str(excel_count)

        for shown_key,shown_item in excel_output_shown_data.items():
            key_cell = 'F'
            item_cell = 'G'
            merge_cell = 'H'
            Rarity_cell = 'I'
            key_cell += str(excel_count)
            item_cell += str(excel_count)
            merge_cell += str(excel_count)
            Rarity_cell += str(excel_count)

            composite_str_cell = 'K'
            composite_str_merge_cell = 'L'
            composite_str_Rarity_cell = 'M'
            composite_str_cell += str(excel_count)
            composite_str_merge_cell += str(excel_count)
            composite_str_Rarity_cell += str(excel_count)

            worksheet.row_dimensions[excel_count].height = 30
            excel_count += 1

            shown_Desc = shown_item["Achievement_Desc"]
            shown_Rarity = checkRarity(shown_item["Rarity"])
            total_Stellar_Jade += shown_Rarity
            composite_str = f"{shown_key}\n{shown_Desc}"

            worksheet[f"{key_cell}"].value = shown_key
            worksheet[f"{key_cell}"].alignment = alignment
            #worksheet.merge_cells(f'{item_cell}:{merge_cell}')
            worksheet[f"{item_cell}"].value = shown_Desc
            worksheet[f"{item_cell}"].alignment = alignment

            worksheet.merge_cells(f'{composite_str_cell}:{composite_str_merge_cell}')
            worksheet[f"{composite_str_cell}"].value = composite_str
            worksheet[f"{composite_str_cell}"].alignment = alignment

            worksheet[f"{Rarity_cell}"].value = shown_Rarity
            worksheet[f"{Rarity_cell}"].alignment = alignment
            worksheet[f"{composite_str_Rarity_cell}"].value = shown_Rarity
            worksheet[f"{composite_str_Rarity_cell}"].alignment = alignment
            
            total_amount += 1

        shown_cell_filling_end = 'I'
        shown_cell_filling_end += str(excel_count-1)
        filling(shown_cell_filling_start,shown_cell_filling_end,worksheet)

        composite_str_filling_end = 'M'
        composite_str_filling_end += str(excel_count-1)
        filling(composite_str_filling_start,composite_str_filling_end,worksheet)

        #輸出普通成就數量
        worksheet['D5'].value = total_amount
        worksheet['D5'].alignment = alignment
        worksheet['C5'].value = "Shown Achievement Amount:"
        worksheet['C5'].alignment = alignment
        
        excel_count += 2
        All_amount = total_amount
        total_amount = 0
        hidden_cell_filling_start = 'F'
        hidden_cell_filling_start += str(excel_count)
        composite_str_filling_start = 'K'
        composite_str_filling_start += str(excel_count)

        for hidden_key,hidden_item in excel_output_hidden_data.items():
            key_cell = 'F'
            item_cell = 'G'
            merge_cell = 'H'
            Rarity_cell = 'I'
            key_cell += str(excel_count)
            item_cell += str(excel_count)
            merge_cell += str(excel_count)
            Rarity_cell += str(excel_count)

            composite_str_cell = 'K'
            composite_str_merge_cell = 'L'
            composite_str_Rarity_cell = 'M'
            composite_str_cell += str(excel_count)
            composite_str_merge_cell += str(excel_count)
            composite_str_Rarity_cell += str(excel_count)

            worksheet.row_dimensions[excel_count].height = 30
            excel_count += 1

            hidden_Desc = hidden_item["Achievement_Desc"]
            hidden_Rarity = checkRarity(hidden_item["Rarity"])
            total_Stellar_Jade += hidden_Rarity
            composite_str = f"{hidden_key}\n{hidden_Desc}"

            worksheet[f"{key_cell}"].value = hidden_key
            worksheet[f"{key_cell}"].alignment = alignment
            #worksheet.merge_cells(f'{item_cell}:{merge_cell}')
            worksheet[f"{item_cell}"].value = hidden_Desc
            worksheet[f"{item_cell}"].alignment = alignment

            worksheet.merge_cells(f'{composite_str_cell}:{composite_str_merge_cell}')
            worksheet[f"{composite_str_cell}"].value = composite_str
            worksheet[f"{composite_str_cell}"].alignment = alignment

            worksheet[f"{Rarity_cell}"].value = hidden_Rarity
            worksheet[f"{Rarity_cell}"].alignment = alignment
            worksheet[f"{composite_str_Rarity_cell}"].value = shown_Rarity
            worksheet[f"{composite_str_Rarity_cell}"].alignment = alignment
            total_amount += 1
        
        hidden_cell_filling_end = 'I'
        hidden_cell_filling_end += str(excel_count-1)
        filling(hidden_cell_filling_start,hidden_cell_filling_end,worksheet)


        composite_str_filling_end = 'M'
        composite_str_filling_end += str(excel_count-1)
        filling(composite_str_filling_start,composite_str_filling_end,worksheet)
        
        All_amount += total_amount
        worksheet['D6'].value = total_amount
        worksheet['D6'].alignment = alignment
        worksheet['C6'].value = "Hidden Achievement Amount:"
        worksheet['C6'].alignment = alignment
        worksheet['D7'].value = All_amount
        worksheet['D7'].alignment = alignment
        worksheet['C7'].value = "ALL Amount:"
        worksheet['C7'].alignment = alignment
        worksheet['D8'].value = total_Stellar_Jade
        worksheet['D8'].alignment = alignment
        worksheet['C8'].value = "Total Stellar Jade:"
        worksheet['C8'].alignment = alignment
        filling('C5','D8',worksheet)
 
    default_sheet = workbook['Sheet']
    workbook.remove(default_sheet)
    workbook.save(new_filename)
    os.system(f"start excel {new_filename}")

def FindTextMap(name, text_map):
    for text, data in text_map.items():
        if str(name) in str(text):
            return data
    return False

def FindSeriesID(Series_ID, series_list):
    for id, data in series_list.items():
        if str(Series_ID) in str(id):
            return data
    return False

def print_clear(strprint):
    print(f"{strprint}", end="\r")
    time.sleep(0.01)

TextMap = { }
AchievementData = { }
AchievementSeries = { }
TextJoinItem = { }
TextJoinConfig = { }
serieslist = { }
excel_output = { }
Series = { }
output_file = 'output.txt'
count_print = 1
print_word = 'Reading '

#下載所需文件
print(f'Downloading data...\n')
for file_name, file_url in Achievement_urls.items():
    response = requests.get(file_url)
    if response.status_code == 200:
        try:
            data = json.loads(response.content.decode('utf-8'))
            if file_name == 'AchievementData':
                AchievementData = data
            elif file_name == 'AchievementSeries':
                AchievementSeries = data
            elif file_name == 'TextJoinItem':
                TextJoinItem = data
            elif file_name == 'TextJoinConfig':
                TextJoinConfig = data
            print(f'{file_name} Download complete\n')
        except json.JSONDecodeError as e:
            print(f"Failed to parse JSON for {file_name}: {e}\n")
    else:
        print(f"{file_name} download failed\n")

print('All data download complete\n')

#讓用戶輸入語言
while True:
    Language = input('Please select language(CHS,CHT,DE,EN,ES,FR,ID,JP,KR,PT,RU,TH,VI) or press Enter to quit:')
    if not Language:
        break
    if Language and Language.upper() in file_urls:
        Language = Language.upper()
        url = file_urls[Language]
        print(f'Downloading {Language}textmap...\n')
        response = requests.get(url)
        if response.status_code == 200:
            TextMap = json.loads(response.content.decode('utf-8'))
            print('TextMap download complete\n')
        else:
            print('Textmap download failed,please retry the progrem.\n')
        #找出所有成就種類
        for series_id, series_data in AchievementSeries.items():
            if "SeriesTitle" in series_data:
                series_title = series_data["SeriesTitle"]
            if "Hash" in series_title:
                hash_value = series_title["Hash"]
                serieslist[series_id] = FindTextMap(hash_value,TextMap)
                excel_output[f"{FindTextMap(hash_value,TextMap)}"] = {"shown":{}, "hidden":{},"hash_value":hash_value}

        #製作excel_output
        print("->")
        for Achievement_id, Achievement_data in AchievementData.items():
            SeriesID = Achievement_data["SeriesID"]
            if "AchievementTitle" in Achievement_data:
                Achievement_Title = Achievement_data["AchievementTitle"]
                AchievementDesc = Achievement_data["AchievementDesc"]["Hash"]
                if "Hash" in Achievement_Title:
                    Rarity = Achievement_data["Rarity"]
                    Ach_hash_value = Achievement_Title["Hash"]
                    Desc = FindTextMap(AchievementDesc,TextMap)
                    Ach_hash_key = FindTextMap(Ach_hash_value,TextMap)
                    
            
                    #刪除常用標籤
                    tags_to_remove = ["</unbreak>", "<unbreak>", "</color>","<color=#8790abff>","<u>","</u>"]
                    for tag in tags_to_remove:
                        Desc = Desc.replace(tag, "")
                        Ach_hash_key = Ach_hash_key.replace(tag, "")

                    #調整換行符號    
                    Desc = Desc.replace("\\n", "\n")
                    Desc = Desc.replace("{TEXTJOIN#54}", FindTextMap(-262052143,TextMap))
                    Desc = Desc.replace("{NICKNAME}", FindTextMap(-2090701432,TextMap))
            

                    #填入#1[i]變數
                    param_list = Achievement_data["ParamList"]
                    param_list_count = 1
                    for param in param_list:
                        param_list_str = '#'
                        value = param["Value"]
                        if isinstance(value, (int, float)):
                            if isinstance(value, float):
                                value = int(value * 100)
                        param_list_str +=  str(param_list_count)
                        param_list_str += '[i]'
                        param_list_str2 = '#'
                        param_list_str2 +=  str(param_list_count)
                        Desc = Desc.replace(param_list_str, str(value))
                        Desc = Desc.replace(param_list_str2, str(value))
                        param_list_count+=1
            
                    #將普通與隱藏成就分別存入
                    Ach_hash_value_str = FindTextMap(Ach_hash_value,TextMap)
                    new_print_word = print_word + str(count_print) + " achievements...."
                    if "ShowType" in Achievement_data:
                        if Achievement_data["ShowType"] == "ShowAfterFinish" :
                            excel_output[f"{FindSeriesID(SeriesID,serieslist)}"]["hidden"][Ach_hash_key] = {"Achievement_Desc":f"{Desc}","Rarity":f"{Rarity}"}
                            #print(excel_output[FindSeriesID(SeriesID,serieslist)]["hidden"][FindTextMap(Ach_hash_value,TextMap)])
                        else:
                            excel_output[f"{FindSeriesID(SeriesID,serieslist)}"]["shown"][Ach_hash_key] = {"Achievement_Desc":f"{Desc}","Rarity":f"{Rarity}"}
                            #print(excel_output[FindSeriesID(SeriesID,serieslist)]["shown"][FindTextMap(Ach_hash_value,TextMap)])
                    else:
                        excel_output[f"{FindSeriesID(SeriesID,serieslist)}"]["shown"][Ach_hash_key] = {"Achievement_Desc":f"{Desc}","Rarity":f"{Rarity}"}
                        #print(excel_output[FindSeriesID(SeriesID,serieslist)]["shown"][FindTextMap(Ach_hash_value,TextMap)])
                    print_clear(new_print_word)
                    count_print += 1
        #輸出excel表
        print_clear("All Achievements read complete！\n\n")

        #json_filename = "output.json"
        #with open(json_filename, 'w', encoding='utf-8') as json_file:
        #    json.dump(excel_output, json_file, ensure_ascii=False, indent=4)
        output_excel(excel_output)
        #output_excel2(excel_output)
    else:
        print('Invalid input\n')


